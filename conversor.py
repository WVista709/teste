import csv
import openpyxl
import os
import xlrd

def agrupar_excels_em_um(arquivos_selecionados, caminho_arquivo_saida):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for aba_nome, caminho_arquivo in arquivos_selecionados.items():
        ws = wb.create_sheet(title=aba_nome)
        ext = os.path.splitext(caminho_arquivo)[1].lower()

        if ext == '.csv':
            with open(caminho_arquivo, encoding='latin1') as f:
                reader = csv.reader(f, delimiter=';')
                primeira_linha = next(reader)
                if primeira_linha and primeira_linha[0].startswith('sep='):
                    primeira_linha = next(reader)
                ws.append(primeira_linha)
                for row in reader:
                    linha_convertida = [converter_valor_csv(celula) for celula in row]
                    ws.append(linha_convertida)

        elif ext == '.xlsx':
            wb_origem = openpyxl.load_workbook(caminho_arquivo, data_only=True)
            ws_origem = wb_origem.active
            for row in ws_origem.iter_rows(values_only=True):
                linha_convertida = [converter_valor(celula) for celula in row]
                ws.append(linha_convertida)

        elif ext == '.xls':
            wb_origem = xlrd.open_workbook(caminho_arquivo)
            ws_origem = wb_origem.sheet_by_index(0)
            for row_idx in range(ws_origem.nrows):
                row = ws_origem.row_values(row_idx)  # Corrigido aqui
                linha_convertida = [converter_valor_csv(celula) for celula in row]
                ws.append(linha_convertida)

        else:
            print(f"Extensão {ext} não suportada para o arquivo {caminho_arquivo}")

    wb.save(caminho_arquivo_saida)
    return caminho_arquivo_saida

def converter_valor_csv(valor):
    if valor is None:
        return None
    valor_str = str(valor).strip()
    try:
        # Remove pontos usados como separador de milhares
        valor_sem_ponto = valor_str.replace('.', '')
        # Substitui vírgula decimal por ponto
        valor_corrigido = valor_sem_ponto.replace(',', '.')
        f = float(valor_corrigido)
        if f.is_integer():
            return int(f)
        else:
            return f
    except ValueError:
        return valor

def converter_valor(valor):
    if valor is None:
        return None
    valor_str = str(valor).strip()
    try:
        valor_corrigido = valor_str.replace(',', '.')
        f = float(valor_corrigido)
        if f.is_integer():
            return int(f)
        else:
            return f
    except ValueError:
        return valor
