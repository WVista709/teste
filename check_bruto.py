from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

class Check:
    @staticmethod
    def formatacao_numeros(ws, min_row, max_row, min_col, max_col):
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row = row, column = col)
                cell.number_format = '#,##0.00;-#,##0.00;-'

    @staticmethod
    def check_compras(caminho_final, modo):

        wb = load_workbook(caminho_final)
        abas_existentes = wb.sheetnames

        if "CHECK" not in abas_existentes:
            ws = wb.create_sheet("CHECK")
        else:
            ws = wb["CHECK"]
        
        #Formatando
        Check.formatacao_numeros(ws, 1, 6, 1, 4)

        # Cabeçalhos
        CelulaValorMesclada(ws, 1, 4, "COMPRAS", linha=1)

        # Linha 2 - Subcabeçalhos
        CelulaValor(ws, 1, "CANCELADAS", linha=2)
        CelulaValor(ws, 2, "SEFAZ", linha=2)
        CelulaValor(ws, 3, "ALTERDATA", linha=2)

        if modo.strip().lower() == "check": 
            CelulaValor(ws, 4, "PRODUTOS", linha=2)
            CelulaValor(ws, 4, "=C5-D5", linha=6)

        # Dados
        CelulaValor(ws, 1, "NÃO", linha=3)
        CelulaValor(ws, 1, "SIM", linha=4)
        CelulaValor(ws, 1, "TOTAL", linha=5)
        CelulaValor(ws, 3, "=B5-C5", linha=6)

        # Fórmulas - só se a aba existir
        if "COMPRAS SEFAZ" in abas_existentes:
            CelulaValor(ws, 2, '=SUMIFS(\'COMPRAS SEFAZ\'!P:P,\'COMPRAS SEFAZ\'!W:W,"NÃO")', linha=3)
            CelulaValor(ws, 2, '=SUMIFS(\'COMPRAS SEFAZ\'!P:P,\'COMPRAS SEFAZ\'!W:W,"SIM")', linha=4)
            CelulaValor(ws, 2, '=B3+B4', linha=5)

        if "COMPRAS ALTERDATA" in abas_existentes:
            CelulaValor(ws, 3, '=SUMIFS(\'COMPRAS ALTERDATA\'!J:J,\'COMPRAS ALTERDATA\'!I:I,"NÃO")', linha=3)
            CelulaValor(ws, 3, '=SUMIFS(\'COMPRAS ALTERDATA\'!J:J,\'COMPRAS ALTERDATA\'!I:I,"SIM")', linha=4)
            CelulaValor(ws, 3, '=C3+C4', linha=5)

        if "COMPRAS PRODUTOS" in abas_existentes:
            CelulaValor(ws, 4, '=SUMIFS(\'COMPRAS PRODUTOS\'!I:I,\'COMPRAS PRODUTOS\'!H:H,"NÃO")', linha=3)
            CelulaValor(ws, 4, '=SUMIFS(\'COMPRAS PRODUTOS\'!I:I,\'COMPRAS PRODUTOS\'!H:H,"SIM")', linha=4)
            CelulaValor(ws, 4, '=D3+D4', linha=5)

        wb.save(caminho_final)

    @staticmethod
    def check_vendas(caminho_final, modo):
        wb = load_workbook(caminho_final)
        abas_existentes = wb.sheetnames

        if "CHECK" not in abas_existentes:
            ws = wb.create_sheet("CHECK")
        else:
            ws = wb["CHECK"]

        #Formatando
        Check.formatacao_numeros(ws, 8, 13, 1, 4)

        # Cabeçalhos
        CelulaValorMesclada(ws, 1, 4, "VENDAS", linha=8)

        # Linha 9 - Subcabeçalhos
        CelulaValor(ws, 1, "CANCELADAS", linha=9)
        CelulaValor(ws, 2, "SEFAZ", linha=9)
        CelulaValor(ws, 3, "ALTERDATA", linha=9)

        if modo.strip().lower() == "check":
            CelulaValor(ws, 4, "PRODUTOS", linha=9)
            CelulaValor(ws, 4, "=C12-D12", linha=13)

        # Dados
        CelulaValor(ws, 1, "NÃO", linha=10)
        CelulaValor(ws, 1, "SIM", linha=11)
        CelulaValor(ws, 1, "TOTAL", linha=12)
        CelulaValor(ws, 3, "=B12-C12", linha=13)

        # Fórmulas - só se as abas existirem
        if "VENDAS SEFAZ" in abas_existentes:
            CelulaValor(ws, 2, '=SUMIFS(\'VENDAS SEFAZ\'!P:P,\'VENDAS SEFAZ\'!W:W,"NÃO")', linha=10)
            CelulaValor(ws, 2, '=SUMIFS(\'VENDAS SEFAZ\'!P:P,\'VENDAS SEFAZ\'!W:W,"SIM")', linha=11)
            CelulaValor(ws, 2, '=B10+B11', linha=12)

        if "VENDAS ALTERDATA" in abas_existentes:
            CelulaValor(ws, 3, '=SUMIFS(\'VENDAS ALTERDATA\'!J:J,\'VENDAS ALTERDATA\'!I:I,"NÃO")', linha=10)
            CelulaValor(ws, 3, '=SUMIFS(\'VENDAS ALTERDATA\'!J:J,\'VENDAS ALTERDATA\'!I:I,"SIM")', linha=11)
            CelulaValor(ws, 3, '=C10+C11', linha=12)

        if "VENDAS PRODUTOS" in abas_existentes:
            CelulaValor(ws, 4, '=SUMIFS(\'VENDAS PRODUTOS\'!I:I,\'VENDAS PRODUTOS\'!H:H,"NÃO")', linha=10)
            CelulaValor(ws, 4, '=SUMIFS(\'VENDAS PRODUTOS\'!I:I,\'VENDAS PRODUTOS\'!H:H,"SIM")', linha=11)
            CelulaValor(ws, 4, '=D10+D11', linha=12)

        wb.save(caminho_final)

def sefaz(caminho_final, modo):
    abas = ["COMPRAS SEFAZ", "VENDAS SEFAZ"]
    wb = load_workbook(caminho_final)
    abas_existentes = wb.sheetnames
    aba_produto = None

    for aba in abas:
        if aba not in abas_existentes:
            continue

        ws = wb[aba]
        colunas, linhas = contar_colunas_linhas_preenchidas(caminho_final, aba)

        # Cabeçalhos
        CelulaValor(ws, colunas + 1, "CANCELADAS", linha=1)
        CelulaValor(ws, colunas + 2, "ALTERDATA", linha=1)
        if modo.strip().lower() == "check":
            if aba == "COMPRAS SEFAZ":
                aba_produto = "COMPRAS PRODUTOS"
            else:
                aba_produto = "VENDAS PRODUTOS"
            CelulaValor(ws, colunas + 3, "PRODUTOS", linha=1)

        for linha in range(2, linhas + 1):
            formula = f'=IF(N{linha}="AUTORIZADA", "NÃO", "SIM")'
            CelulaValor(ws, colunas + 1, formula, linha=linha)

        # Define a aba de referência
        if aba == "COMPRAS SEFAZ":
            aba_referencia = "COMPRAS ALTERDATA"
        else:  # VENDAS SEFAZ
            aba_referencia = "VENDAS ALTERDATA"

        # Insere fórmula para aba_referencia se existir
        if aba_referencia in abas_existentes:
            for linha in range(2, linhas + 1):
                formula = f'=IFERROR(VLOOKUP(C{linha},\'{aba_referencia}\'!B:B,1,0),"ERRO")'
                CelulaValor(ws, colunas + 2, formula, linha=linha)
        else:
            print(f"Aba de referência '{aba_referencia}' não existe. Fórmulas não inseridas em '{aba}'.")

        # Insere fórmula para aba_produto se modo Check e aba_produto existir
        if modo.strip().lower() == "check" and aba_produto in abas_existentes:
            for linha in range(2, linhas + 1):
                formula = f'=IFERROR(VLOOKUP(C{linha},\'{aba_produto}\'!B:B,1,0),"ERRO")'
                CelulaValor(ws, colunas + 3, formula, linha=linha)
        elif modo.strip().lower() == "check":
            print(f"Aba de referência '{aba_produto}' não existe. Fórmulas PRODUTOS não inseridas em '{aba}'.")

    wb.save(caminho_final)

def alterdata(caminho_final, modo):
    abas = ["COMPRAS ALTERDATA", "VENDAS ALTERDATA"]
    wb = load_workbook(caminho_final)
    abas_existentes = wb.sheetnames
    aba_produto = None

    for aba in abas:
        if aba not in abas_existentes:
            continue

        ws = wb[aba]
        colunas, linhas = contar_colunas_linhas_preenchidas(caminho_final, aba)
        CelulaValor(ws, colunas + 1, "SEFAZ", linha=1)

        if modo.strip().lower() == "check":
            if aba == "COMPRAS ALTERDATA":
                aba_produto = "COMPRAS PRODUTOS"
            else:  # VENDAS ALTERDATA
                aba_produto = "VENDAS PRODUTOS"
            CelulaValor(ws, colunas + 2, "PRODUTOS", linha=1)

        # Define a aba de referência
        if aba == "COMPRAS ALTERDATA":
            aba_referencia = "COMPRAS SEFAZ"
        else:  # VENDAS ALTERDATA
            aba_referencia = "VENDAS SEFAZ"

        # Insere fórmula para aba_referencia se existir
        if aba_referencia in abas_existentes:
            for linha in range(2, linhas + 1):
                formula = f'=IFERROR(VLOOKUP(B{linha},\'{aba_referencia}\'!C:C,1,0),"ERRO")'
                CelulaValor(ws, colunas + 1, formula, linha=linha)
        else:
            print(f"Aba de referência '{aba_referencia}' não existe. Fórmulas não inseridas em '{aba}'.")

        # Insere fórmula para aba_produto se modo Check e aba_produto existir
        if modo.strip().lower() == "check" and aba_produto in abas_existentes:
            for linha in range(2, linhas + 1):
                formula = f'=IFERROR(VLOOKUP(B{linha},\'{aba_produto}\'!B:B,1,0),"ERRO")'
                CelulaValor(ws, colunas + 2, formula, linha=linha)
        elif modo.strip().lower() == "check":
            print(f"Aba de referência '{aba_produto}' não existe. Fórmulas PRODUTOS não inseridas em '{aba}'.")

    wb.save(caminho_final)

def produto(caminho_final):
    abas = ["COMPRAS PRODUTOS", "VENDAS PRODUTOS"]
    wb = load_workbook(caminho_final)
    abas_existentes = wb.sheetnames

    for aba in abas:
        if aba not in abas_existentes:
            continue

        ws = wb[aba]
        colunas, linhas = contar_colunas_linhas_preenchidas(caminho_final, aba)
        CelulaValor(ws, colunas + 1, "SEFAZ", linha=1)
        CelulaValor(ws, colunas + 2, "ALTERDATA", linha=1)

        # Define as abas de referência para SEFAZ e ALTERDATA
        if aba == "COMPRAS PRODUTOS":
            aba_sefaz = "COMPRAS SEFAZ"
            aba_alterdata = "COMPRAS ALTERDATA"
        else:  # VENDAS PRODUTOS
            aba_sefaz = "VENDAS SEFAZ"
            aba_alterdata = "VENDAS ALTERDATA"

        # Insere fórmula para SEFAZ se existir
        if aba_sefaz in abas_existentes:
            for linha in range(2, linhas + 1):
                formula_sefaz = f'=IFERROR(VLOOKUP(B{linha},\'{aba_sefaz}\'!C:C,1,0),"ERRO")'
                CelulaValor(ws, colunas + 1, formula_sefaz, linha=linha)
        else:
            print(f"Aba de referência '{aba_sefaz}' não existe. Fórmulas SEFAZ não inseridas em '{aba}'.")

        # Insere fórmula para ALTERDATA se existir
        if aba_alterdata in abas_existentes:
            for linha in range(2, linhas + 1):
                formula_alterdata = f'=IFERROR(VLOOKUP(B{linha},\'{aba_alterdata}\'!B:B,1,0),"ERRO")'
                CelulaValor(ws, colunas + 2, formula_alterdata, linha=linha)
        else:
            print(f"Aba de referência '{aba_alterdata}' não existe. Fórmulas ALTERDATA não inseridas em '{aba}'.")

    wb.save(caminho_final)

def contar_colunas_linhas_preenchidas(caminho_arquivo, nome_aba):
    wb = load_workbook(caminho_arquivo, data_only=True)
    if nome_aba not in wb.sheetnames:
        return 0, 0
    ws = wb[nome_aba]
    primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    colunas_preenchidas = sum(1 for cell in primeira_linha if cell not in (None, ""))
    linhas_preenchidas = sum(
        1 for row in ws.iter_rows(values_only=True)
        if any(cell not in (None, "") for cell in row)
    )
    return colunas_preenchidas, linhas_preenchidas

def CelulaValor(ws, coluna_num, valor, linha=1):
    letra_coluna = get_column_letter(coluna_num)
    ws[f"{letra_coluna}{linha}"] = valor

def CelulaValorMesclada(ws, coluna_inicio, coluna_fim, valor, linha=1):
    letra_inicio = get_column_letter(coluna_inicio)
    letra_fim = get_column_letter(coluna_fim)
    celula_inicio = f"{letra_inicio}{linha}"
    celula_fim = f"{letra_fim}{linha}"
    ws.merge_cells(f"{celula_inicio}:{celula_fim}")
    ws[celula_inicio] = valor